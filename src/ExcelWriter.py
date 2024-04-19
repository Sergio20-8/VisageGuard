from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import datetime

def escribir_en_excel(diccionario, nombre_archivo):
    try:
        # Cargar el libro de trabajo existente si existe
        wb = load_workbook(filename=nombre_archivo)
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        wb = Workbook()
        ws = wb.active
        # Agregar encabezados a las columnas si es un archivo nuevo
        ws.append(["Nombre", datetime.datetime.now().strftime("%m/%d/%y")])

    # Obtener la fecha y hora actual
    fecha_actual = datetime.datetime.now().strftime("%m/%d/%y")
    hora_actual = datetime.datetime.now().strftime("%I:%M %p")  # Formato AM/PM

    # Verificar si la fecha actual ya está presente como encabezado de columna
    if fecha_actual not in [cell.value for cell in ws[1]]:
        # Si no está presente, agregarla como un nuevo encabezado de columna
        ws.cell(row=1, column=len(ws[1]) + 1, value=fecha_actual)

    # Obtener el índice de la columna para la fecha actual
    column_index = [cell.value for cell in ws[1]].index(fecha_actual) + 1

    # Obtener la última fila ocupada en la hoja de cálculo
    last_row = ws.max_row

    # Agregar datos del diccionario a las filas
    for carnet, info in diccionario.items():
        # Verificar si el alumno ha sido reconocido
        if info["reconocido"]:
            # Obtener el índice de la fila para el nombre del alumno
            if carnet not in [cell.value for cell in ws['A']]:
                ws.append([carnet])  # Si el nombre no está presente, agregarlo en una nueva fila
            row_index = [cell.value for cell in ws['A']].index(carnet) + 1

            # Escribir la hora de detección en la celda correspondiente
            ws.cell(row=row_index, column=column_index, value=hora_actual)
            ws.cell(row=row_index, column=column_index).alignment = Alignment(horizontal='center')
            # Cambiar formato de celda a verde para las personas que asistieron
            ws.cell(row=row_index, column=column_index).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
        else:
            # Si el alumno no fue reconocido, escribir su nombre y "No asistió" debajo de la fecha actual
            if carnet not in [cell.value for cell in ws['A']]:
                ws.append([carnet])  # Si el nombre no está presente, agregarlo en una nueva fila
            row_index = [cell.value for cell in ws['A']].index(carnet) + 1
            ws.cell(row=row_index, column=column_index, value="No asistió")
            ws.cell(row=row_index, column=column_index).alignment = Alignment(horizontal='center')
            # Cambiar formato de celda a rojo con texto blanco y negrita
            ws.cell(row=row_index, column=column_index).font = Font(color="FFFFFF", bold=True)
            ws.cell(row=row_index, column=column_index).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Guardar el libro de trabajo
    wb.save(nombre_archivo)
    print(f"Archivo Excel '{nombre_archivo}' actualizado exitosamente.")
